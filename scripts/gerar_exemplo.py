#!/usr/bin/env python3
"""Gera uma planilha de exemplo com a estrutura descrita em config/fontes.yml.

Serve para dois propósitos:

1. rodar o pipeline de ponta a ponta antes de o xlsm real existir;
2. exercitar o validador — o exemplo inclui de propósito os defeitos típicos de
   planilha preenchida à mão (espaço sobrando, número em formato brasileiro,
   data como texto, categoria fora da lista, município inexistente na
   tabela de apoio, nota faltando).

Uso:  python scripts/gerar_exemplo.py [--linhas 120] [--saida data/raw/...]

O arquivo gerado é ignorado pelo git (ver .gitignore: data/raw/exemplo_*).
"""

from __future__ import annotations

import argparse
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

RAIZ = Path(__file__).resolve().parents[1]

TERRITORIOS = {
    "Salvador": "Metropolitano de Salvador",
    "Camaçari": "Metropolitano de Salvador",
    "Feira de Santana": "Portal do Sertão",
    "Vitória da Conquista": "Sudoeste Baiano",
    "Itabuna": "Litoral Sul",
    "Ilhéus": "Litoral Sul",
    "Juazeiro": "Sertão do São Francisco",
    "Barreiras": "Bacia do Rio Grande",
    "Jequié": "Médio Sudoeste da Bahia",
    "Paulo Afonso": "Itaparica",
    "Teixeira de Freitas": "Extremo Sul",
    "Irecê": "Irecê",
    "Santo Antônio de Jesus": "Recôncavo",
    "Alagoinhas": "Litoral Norte e Agreste Baiano",
    "Porto Seguro": "Costa do Descobrimento",
}
POPULACAO = {
    "Salvador": 2417678,
    "Camaçari": 300660,
    "Feira de Santana": 616279,
    "Vitória da Conquista": 370938,
    "Itabuna": 213685,
    "Ilhéus": 159362,
    "Juazeiro": 218162,
    "Barreiras": 154610,
    "Jequié": 158459,
    "Paulo Afonso": 117437,
    "Teixeira de Freitas": 163449,
    "Irecê": 78544,
    "Santo Antônio de Jesus": 105583,
    "Alagoinhas": 155747,
    "Porto Seguro": 148686,
}
EIXOS = [
    "Educação",
    "Geração de renda",
    "Cultura e identidade",
    "Direitos e cidadania",
    "Meio ambiente",
    "Saúde comunitária",
]
ETAPAS = ["Documentação", "Análise técnica", "Visita de campo", "Comitê", "Concluída"]
STATUS = [
    "Inscrita",
    "Em análise",
    "Habilitada",
    "Inabilitada",
    "Selecionada",
    "Não selecionada",
    "Desistente",
]
CRITERIOS = [
    "Relevância territorial",
    "Capacidade institucional",
    "Coerência do orçamento",
    "Potencial de rede",
    "Sustentabilidade",
]


def gerar(linhas: int, saida: Path, semente: int = 7) -> Path:
    rnd = random.Random(semente)
    wb = Workbook()

    # --- Inscricoes -----------------------------------------------------------
    ws = wb.active
    ws.title = "Inscricoes"
    ws.append(
        [
            "ID Inscrição",
            "Organização",
            "CNPJ",
            "E-mail de Contato",
            "Município",
            "Território de Identidade",
            "Eixo",
            "Etapa",
            "Status",
            "Data de Inscrição",
            "Valor Solicitado",
            "Nota Final",
        ]
    )
    base = date(2026, 3, 2)
    municipios = list(TERRITORIOS)
    ids: list[str] = []

    for i in range(1, linhas + 1):
        codigo = f"RB-2026-{i:04d}"
        ids.append(codigo)
        municipio = rnd.choice(municipios)
        status = rnd.choices(STATUS, weights=[18, 24, 16, 8, 12, 16, 6])[0]
        avaliada = status in ("Habilitada", "Selecionada", "Não selecionada")
        dia = base + timedelta(days=rnd.randint(0, 110))
        valor = round(rnd.uniform(25_000, 320_000), 2)

        # Defeitos plantados de propósito (ver docstring).
        if i == 3:
            municipio_celula = f"  {municipio}  "  # espaço sobrando
        elif i == 7:
            municipio_celula = "Bom Jesus da Lapa"  # fora da tabela de apoio
        else:
            municipio_celula = municipio
        valor_celula = (
            f"R$ {valor:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
            if i % 17 == 0
            else valor
        )  # número em formato brasileiro
        data_celula = dia.strftime("%d/%m/%Y") if i % 13 == 0 else dia  # data como texto
        status_celula = "Em Análise " if i == 11 else status  # categoria fora da lista

        cnpj = (
            f"{rnd.randint(10, 99)}.{rnd.randint(100, 999)}."
            f"{rnd.randint(100, 999)}/0001-{rnd.randint(10, 99)}"
        )
        ws.append(
            [
                codigo,
                f"Associação Rede {i:03d}" if i % 2 else f"Instituto Redes {i:03d}",
                cnpj,
                f"contato{i:03d}@exemplo.org.br",
                municipio_celula,
                TERRITORIOS[municipio],
                rnd.choice(EIXOS),
                rnd.choice(ETAPAS),
                status_celula,
                data_celula,
                valor_celula,
                round(rnd.uniform(4.5, 9.8), 1) if avaliada else None,
            ]
        )

    # --- Avaliacoes -----------------------------------------------------------
    wa = wb.create_sheet("Avaliacoes")
    wa.append(["ID Inscrição", "Avaliador", "Critério", "Nota", "Data da Avaliação"])
    for codigo in ids:
        if rnd.random() < 0.45:
            continue
        for avaliador in rnd.sample(["AV-01", "AV-02", "AV-03", "AV-04"], k=2):
            for criterio in CRITERIOS:
                wa.append(
                    [
                        codigo,
                        avaliador,
                        criterio,
                        round(rnd.uniform(3.0, 10.0), 1),
                        base + timedelta(days=rnd.randint(30, 140)),
                    ]
                )
    wa.append(["RB-2026-9999", "AV-01", CRITERIOS[0], 8.0, base])  # referência órfã

    # --- Municipios -----------------------------------------------------------
    wm = wb.create_sheet("Municipios")
    wm.append(["Município", "Código IBGE", "Território de Identidade", "População"])
    for j, (municipio, territorio) in enumerate(TERRITORIOS.items(), start=1):
        wm.append([municipio, f"29{j:05d}", territorio, POPULACAO[municipio]])

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


def main() -> int:
    hoje = date.today().isoformat()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--linhas", type=int, default=120, help="número de inscrições")
    parser.add_argument(
        "--saida",
        default=RAIZ / "data" / "raw" / f"exemplo_{hoje}_redes_bahia.xlsm",
        type=Path,
    )
    args = parser.parse_args()
    destino = gerar(args.linhas, args.saida)
    print(f"Planilha de exemplo criada em {destino}")
    print("Rode agora: make dados")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
